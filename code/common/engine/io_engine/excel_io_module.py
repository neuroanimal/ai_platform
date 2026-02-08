"""
Excel IO Module - XLS/XLSX processing with table extraction
Optimal integration from additional_codes_part_1
"""
import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, List, Optional, Tuple
import warnings

from common.handler.trace_handler import TraceHandler
from common.handler.error_handler import ErrorHandler, FormatProcessingError


class ExcelIOModule:
    """Excel I/O with table detection and extraction"""

    def __init__(self, tracer: TraceHandler):
        self.tracer = tracer
        self.stats = {
            "files_read": 0,
            "files_written": 0,
            "tables_extracted": 0,
            "sheets_processed": 0
        }

    def read_file(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Read Excel file and return data"""
        try:
            if not os.path.exists(file_path):
                raise FormatProcessingError(f"Excel file not found: {file_path}")

            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning)

                if sheet_name:
                    data = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                    result = {sheet_name: data}
                else:
                    data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                    result = data

            self.stats["files_read"] += 1
            self.stats["sheets_processed"] += len(result)
            self.tracer.info(f"Successfully read Excel file: {file_path}")
            return result

        except Exception as e:
            error_msg = f"Failed to read Excel file {file_path}: {str(e)}"
            self.tracer.error(error_msg)
            raise FormatProcessingError(error_msg)

    def write_file(self, data: Dict[str, pd.DataFrame], output_path: str) -> bool:
        """Write data to Excel file"""
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            self.stats["files_written"] += 1
            self.tracer.info(f"Successfully wrote Excel file: {output_path}")
            return True

        except Exception as e:
            error_msg = f"Failed to write Excel file {output_path}: {str(e)}"
            self.tracer.error(error_msg)
            raise FormatProcessingError(error_msg)

    def extract_tables(self, file_path: str, output_dir: str) -> Dict[str, List[str]]:
        """Extract tables from Excel file to CSV files"""
        try:
            workbook = load_workbook(file_path)
            results = {}

            for worksheet in workbook.worksheets:
                sheet_name = worksheet.title
                self.tracer.info(f"Processing worksheet: {sheet_name}")

                if self._is_sheet_empty(worksheet):
                    self.tracer.warning(f"Worksheet '{sheet_name}' is empty")
                    continue

                # Convert to DataFrame for processing
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)

                if not data:
                    continue

                df = pd.DataFrame(data)
                tables = self._detect_tables(df)

                csv_files = []
                for i, (start_row, end_row, headers) in enumerate(tables):
                    table_df = df.iloc[start_row:end_row+1].copy()
                    table_df.columns = headers

                    # Generate CSV filename
                    csv_filename = f"{sheet_name}_table_{i}_{len(headers)}x{len(table_df)}.csv"
                    csv_path = os.path.join(output_dir, csv_filename)

                    # Write to CSV
                    os.makedirs(output_dir, exist_ok=True)
                    table_df.to_csv(csv_path, sep='\t', index=False, encoding='utf-8')
                    csv_files.append(csv_path)

                    self.stats["tables_extracted"] += 1

                results[sheet_name] = csv_files

            self.tracer.info(f"Extracted {self.stats['tables_extracted']} tables from {file_path}")
            return results

        except Exception as e:
            error_msg = f"Failed to extract tables from {file_path}: {str(e)}"
            self.tracer.error(error_msg)
            raise FormatProcessingError(error_msg)

    def _is_sheet_empty(self, worksheet) -> bool:
        """Check if worksheet is empty"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    return False
        return True

    def _detect_tables(self, df: pd.DataFrame) -> List[Tuple[int, int, List[str]]]:
        """Detect tables in DataFrame"""
        tables = []
        num_rows, num_cols = df.shape

        current_row = 0
        while current_row < num_rows:
            # Look for header row (row with multiple non-null values)
            header_row = -1
            for row_idx in range(current_row, num_rows):
                row = df.iloc[row_idx]
                non_null_count = row.count()

                if non_null_count > 1:  # Found potential header
                    header_row = row_idx
                    break

            if header_row == -1:
                break

            # Extract headers
            headers = [str(val) if val is not None else f"Col_{i}"
                      for i, val in enumerate(df.iloc[header_row])]

            # Find end of table
            end_row = num_rows - 1
            for row_idx in range(header_row + 1, num_rows):
                row = df.iloc[row_idx]
                non_null_count = row.count()

                if non_null_count <= 1:  # End of table
                    end_row = row_idx - 1
                    break

            if end_row > header_row:
                tables.append((header_row + 1, end_row, headers))

            current_row = end_row + 1

        return tables

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names from Excel file"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            self.tracer.error(f"Failed to get sheet names from {file_path}: {str(e)}")
            return []

    def validate_excel(self, file_path: str) -> bool:
        """Validate Excel file"""
        try:
            load_workbook(file_path, read_only=True)
            return True
        except Exception:
            return False

    def get_summary(self) -> Dict[str, Any]:
        """Get processing summary"""
        return self.stats